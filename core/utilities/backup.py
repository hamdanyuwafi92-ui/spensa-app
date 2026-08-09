import io
from django.contrib.auth import get_user_model
from django.db import transaction
from django.db.models import ManyToManyField, FileField, ImageField
from openpyxl import Workbook, load_workbook

User = get_user_model()

EXCLUDED_FIELDS = {"id", "created_at", "updated_at", "created_by", "updated_by"}


def backup_model_to_excel(model_class, extra_fields=None):
    wb = Workbook()
    ws = wb.active
    ws.title = model_class.__name__

    fields = [
        f
        for f in model_class._meta.get_fields()
        if hasattr(f, "column")
        and f.attname not in EXCLUDED_FIELDS
        and not isinstance(f, ManyToManyField)
    ]
    headers = [f.attname for f in fields]
    if extra_fields:
        headers += extra_fields
    ws.append(headers)

    for obj in model_class.objects.all():
        row = []
        for field in fields:
            value = getattr(obj, field.attname)
            if hasattr(value, "pk"):
                value = value.pk
            elif isinstance(field, (FileField, ImageField)):
                value = value.name if value else None
            row.append(value)
        if extra_fields:
            for extra in extra_fields:
                value = getattr(obj, extra, None) if hasattr(obj, extra) else ""
                row.append(value)
        ws.append(row)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def restore_model_from_excel(model_class, file, user=None):
    wb = load_workbook(file)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    instances = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        data.pop("id", None)
        for field in model_class._meta.get_fields():
            if isinstance(field, (FileField, ImageField)) and field.attname in data:
                del data[field.attname]
        data = {k: v for k, v in data.items() if v is not None}
        if user and hasattr(model_class, "created_by"):
            data["created_by"] = user
        instances.append(model_class(**data))

    with transaction.atomic():
        model_class.objects.all().delete()
        model_class.objects.bulk_create(instances)


def restore_teacher_from_excel(file, request_user=None):
    from core.models import Teacher

    wb = load_workbook(file)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    user_list = []
    teacher_list = []

    from django.contrib.auth.hashers import make_password

    hashed_password = make_password("guruoke123")

    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        nip = data.get("nip")
        first_name = data.get("first_name", "")
        last_name = data.get("last_name", "")
        email = data.get("email", "")

        if nip:
            user = User(
                username=nip,
                first_name=first_name,
                last_name=last_name,
                email=email,
                is_active=True,
                password=hashed_password,
            )
            job = data.get("job")
            if job == "Developer":
                user.is_staff = True
                user.is_superuser = True
            elif job in ["Administrator", "Agent", "Guru"]:
                user.is_staff = True
            user_list.append(user)

            teacher_list.append(
                Teacher(
                    user=user,
                    nip=nip,
                    job=job,
                    gender=data.get("gender"),
                    nuptk=data.get("nuptk", ""),
                    gelar_depan=data.get("gelar_depan", ""),
                    gelar_belakang=data.get("gelar_belakang", ""),
                    tempat_lahir=data.get("tempat_lahir", ""),
                    tanggal_lahir=data.get("tanggal_lahir"),
                    alamat=data.get("alamat", ""),
                    nomor_hp=data.get("nomor_hp", ""),
                    created_by=request_user if request_user else None,
                )
            )

    with transaction.atomic():
        Teacher.objects.all().delete()
        User.objects.bulk_create(user_list)
        Teacher.objects.bulk_create(teacher_list)


def restore_student_from_excel(file, request_user=None):
    from core.models import Student

    wb = load_workbook(file)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    user_list = []
    student_list = []

    from django.contrib.auth.hashers import make_password

    hashed_password = make_password("siswaoke123")

    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        nisn = data.get("nisn")
        first_name = data.get("first_name", "")
        last_name = data.get("last_name", "")
        email = data.get("email", "")

        if nisn:
            user = User(
                username=nisn,
                first_name=first_name,
                last_name=last_name,
                email=email,
                is_active=True,
                password=hashed_password,
            )
            user_list.append(user)

            student_list.append(
                Student(
                    user=user,
                    nisn=nisn,
                    job="Siswa",
                    gender=data.get("gender"),
                    status=data.get("status", "Aktif"),
                    nis=data.get("nis", ""),
                    nama_ayah=data.get("nama_ayah", ""),
                    nama_ibu=data.get("nama_ibu", ""),
                    nomor_hp_ortu=data.get("nomor_hp_ortu", ""),
                    tempat_lahir=data.get("tempat_lahir", ""),
                    tanggal_lahir=data.get("tanggal_lahir"),
                    alamat=data.get("alamat", ""),
                    nomor_hp=data.get("nomor_hp", ""),
                    created_by=request_user if request_user else None,
                )
            )

    with transaction.atomic():
        Student.objects.all().delete()
        User.objects.bulk_create(user_list)
        Student.objects.bulk_create(student_list)
